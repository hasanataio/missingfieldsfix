import pandas as pd
from openpyxl import Workbook, load_workbook
import numpy as np


missing_fields={'Menu': {'id': 1,
  'menuName': "Main Menu",
  'posDisplayName': "Main Menu",
  'menuDescription': "Main Menu",
  'restaurantId': 1,
  'sortOrder': 1,
  'posButtonColor': "#e34032"},
 'Category': {'id': 7,
  'posDisplayName': 7,
  'kdsDisplayName': 7,
  'sortOrder': 7,
  'menuIds': 7},
 'Category Items': {'id': 7, 'sortOrder': 7},
 'Item': {'id': 7,
  'showOnMenu': 7,
  'showOnline': 7,
  'showPOS': 7,
  'showQR': 7,
  'showThirdParty': 7,
  'posDisplayName': 7,
  'kdsDisplayName': 7,
  'orderQuantityLimit': 7,
  'minLimit': 7,
  'maxLimit': 7,
  'noMaxLimit': 7},
 'Item Modifiers': {'sortOrder': 7},
 'Modifier': {'id': 7,
  'posDisplayName': 7,
  'multiSelect': 7,
  'isNested': 7,
  'isOptional': 7,
  'priceType': 7,
  'canGuestSelectMoreModifiers': 7,
  'minSelector': 7,
  'maxSelector': 7,
  'isSizeModifier': 7,
  'showOnPos': 7,
  'showOnKiosk': 7,
  'showOnMpos': 7,
  'showOnQR': 7,
  'showOnline': 7,
  'showOnThirdParty': 7,
  'limitIndividualModifierSelection': 7},
 'Modifier Option': {'id': 7,
  'posDisplayName': 7,
  'kdsDisplayName': 7,
  'price': 7,
  'isStockAvailable': 7,
  'isSizeModifier': 7},
 'Modifier ModifierOptions': {'isDefaultSelected': 7, 'maxLimit': 7},
 'Modifier Group': {},
 'Setting': {},
 'Visibility Setting': {},
 'Day Schedule': {},
 'Category ModifierGroups': {},
 'Category Modifiers': {},
 'Allergen': {},
 'Tag': {},
 'Item Modifier Group': {}}


def fix_missing_fields(filename):
    
    dataframes={}
    def read_or_create_sheet(filename, sheet_name):
        try:
            # Try to read the sheet
            data = pd.read_excel(filename, sheet_name=sheet_name)
        except ValueError as e:
            if "not found" in str(e) or "not exist" in str(e):
                # Create a new workbook if the file doesn't exist, otherwise load the existing one
                try:
                    wb = load_workbook(filename)
                except FileNotFoundError:
                    wb = Workbook()
                    wb.save(filename)

                # Add the new sheet
                wb.create_sheet(sheet_name)
                wb.save(filename)

                # Read the newly created sheet
                data = pd.DataFrame()
            else:
                raise e
        return data


    def fix_menu_sheet(cols,sheetname):
        data=read_or_create_sheet(filename,sheet_name=sheetname)
        if len(data['menuName'].values)<1 or np.isnan(list(data['menuName'].values)).any():
            for col in missing_fields[sheetname].keys():
                data[col]=[missing_fields[sheetname][col]]
            dataframes[sheetname]=data

        
    def fix_category_sheet(cols,sheetname):
        data=read_or_create_sheet(filename,sheet_name=sheetname)
        category_names=list(data['categoryName'])
        data['id']=[i for i in range(1,len(data)+1)]
        data['posDisplayName']=category_names
        data['kdsDisplayName']=category_names
        data['sortOrder']=[i for i in range(1,len(data)+1)]
        data['menuIds']=[1 for i in range(1,len(data)+1)]
        dataframes[sheetname]=data

    def fix_category_items_sheet(cols,sheetname):
        
        
        data=read_or_create_sheet(filename,sheet_name=sheetname)
        data['id']=[i for i in range(1,len(data)+1)]
        data['sortOrder']=[i for i in range(1,len(data)+1)]
        dataframes[sheetname]=data

    def fix_items_sheet(cols,sheetname):
        
        data=read_or_create_sheet(filename,sheet_name=sheetname)
        data['id']=[i for i in range(1,len(data)+1)]
        data['posDisplayName']=data['itemName']
        data['kdsDisplayName']=data['itemName']
        data['showOnMenu']=[True for i in range(len(data))]				
        data['showOnline']=[True for i in range(len(data))]	
        data['showPOS']=[True for i in range(len(data))]	
        data['showQR']=[True for i in range(len(data))]	
        data['showThirdParty']=[True for i in range(len(data))]	
        data['orderQuantityLimit']=[True for i in range(len(data))]	
        data['minLimit']=[0 for i in range(len(data))]	
        data['maxLimit']=[50 for i in range(len(data))]
        dataframes[sheetname]=data

    def fix_item_modifiers(cols,sheetname):
        data=read_or_create_sheet(filename,sheet_name=sheetname)
        data['sortOrder']=[i for i in range(1,len(data)+1)]	
        dataframes[sheetname]=data

    def fix_modifier_options(cols,sheetname):
        data=read_or_create_sheet(filename,sheet_name=sheetname)
        data['posDisplayName']=data['optionName']
        data['kdsDisplayName']=data['optionName']
        data['price']=[0 for i in range(len(data))]
        data['isStockAvailable']=[True for i in range(len(data))]
        data['isSizeModifier']=[False for i in range(len(data))]
        dataframes[sheetname]=data

    def modifier_modifider_options(cols,sheetname):
        data=read_or_create_sheet(filename,sheet_name=sheetname)
        data['isDefaultSelected']=[False for i in range(len(data))]
        data['maxLimit']=[1 for i in range(len(data))]
        dataframes[sheetname]=data

    def fix_modifier(cols,sheetname):
        data=read_or_create_sheet(filename,sheet_name=sheetname)
        data['posDisplayName']=data['modifierName']
        data['multiSelect']=[False for i in range(len(data))]
        data['isNested']=[False for i in range(len(data))]
        data['isOptional']=[True for i in range(len(data))]
        data['priceType']=["individual" for i in range(len(data))]
        data['canGuestSelectMoreModifiers']=[True for i in range(len(data))]
        data['minSelector']=[0 for i in range(len(data))]
        data['isSizeModifier']=[False for i in range(len(data))]
        data['showOnPos']=[True for i in range(len(data))]
        data['showOnKiosk']=[True for i in range(len(data))]
        data['showOnMpos']=[True for i in range(len(data))]
        data['showOnQR']=[True for i in range(len(data))]
        data['showOnline']=[True for i in range(len(data))]
        data['showOnThirdParty']=[True for i in range(len(data))]
        data['limitIndividualModifierSelection']=[True for i in range(len(data))]
        df2=pd.read_excel(filename,sheet_name="Modifier ModifierOptions")
        modifier_counts = df2.groupby('modifierId').size().reset_index(name='count')
        data = data.merge(modifier_counts, left_on='id', right_on='modifierId', how='left')
        data['maxSelector'] = data['count'].fillna(0)
        data.drop('count', axis=1, inplace=True)
        data.drop('modifierId', axis=1, inplace=True)
        # data = data.loc[:, ~data.columns.str.contains('^Unnamed')]
        dataframes[sheetname]=data



    def add_remaining(cols,sheetname):
        data=read_or_create_sheet(filename,sheet_name=sheetname)
        dataframes[sheetname]=data

    for sheetname in missing_fields.keys():
        if sheetname=="Menu":
            cols=missing_fields[sheetname].keys()
            fix_menu_sheet(cols,sheetname)
            
                
        elif sheetname=="Category":
            cols=missing_fields[sheetname].keys()
            dataframe=fix_category_sheet(cols,sheetname)
            

        elif sheetname=="Category Items":
            cols=missing_fields[sheetname].keys()
            fix_category_items_sheet(cols,sheetname)

        elif sheetname=="Item":
            cols=missing_fields[sheetname].keys()
            fix_items_sheet(cols,sheetname)

        elif sheetname=="Item Modifiers":
            cols=missing_fields[sheetname].keys()
            fix_item_modifiers(cols,sheetname)
        
        elif sheetname=="Modifier":
            cols=missing_fields[sheetname].keys()
            fix_modifier(cols,sheetname)

        elif sheetname=="Modifier ModifierOptions":
            cols=missing_fields[sheetname].keys()
            modifier_modifider_options(cols,sheetname)

        elif sheetname=="Modifier Option":
            cols=missing_fields[sheetname].keys()
            fix_modifier_options(cols,sheetname)

        else:
            cols=missing_fields[sheetname].keys()
            add_remaining(cols,sheetname)
            

    return dataframes


